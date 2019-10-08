import openpyxl as op
from copy import copy
import re
Hyper_list=[]
wb= op.load_workbook(r'C:\Users\sahil.rana\Desktop\task\Quantum-II-MIM-IA.xlsx')
Inbound=wb.get_sheet_by_name('Inbound')
Outbound=wb.get_sheet_by_name('Outbound')
Index=wb.get_sheet_by_name('Index')
Vendor_sub=[]

def rep2ndrow(job):
    print("inside rep2nd")
    with open(r'C:\Users\sahil.rana\Desktop\task\Checked in R1 inbound\Checked in R1 inbound\{}\MIM\SQL\insert-JOB-DEV-SYS-INT-PRD.sql'.format(job),'r') as obj:
        for i in obj:
            if "??????????????" in i:
                break
    file=i.strip()
    file=file.replace("'",'')
    file=file.replace(",",'')
    return file

def PC_FILEID(job):
    Fid=''
    Vendor=['DMT','INTERAC','TSYS','BMO','TELUS']
    with open(r'C:\Users\sahil.rana\Desktop\task\Checked in R1 inbound\Checked in R1 inbound\{}\MIM\SQL\insert-JOB-DEV-SYS-INT-PRD.sql'.format(job),'r') as obj:
        for i in obj:
            if re.search("values\s\(\d\d\d\d\s+", i):
                Fid=i.strip().split(' ')[1].replace('(','F0')
            for v in Vendor:
                if v in i :
                    global Vendor_sub
                    Vendor_sub=i.strip().replace(',\'','').replace('\'','').split(' ')
                    break
    Fid = Fid[:5]
    return Fid

def edit_Inbound(j):
#    print("inside edit inbound")
    mcol=Inbound.max_column
    for i in range(1,Inbound.max_row+1):
        Inbound.cell(row=i,column=mcol+1).value=Inbound.cell(row=i,column=mcol).value
        Inbound.cell(row=i,column=mcol+1)._style=copy(Inbound.cell(row=i,column=mcol)._style)
        if i==2:
            Inbound.cell(row=i,column=mcol+1).value=rep2ndrow(j.rstrip())
        elif i==4:
            Inbound.cell(row=i,column=mcol+1).value=j
            global Hyper_list
            print(Index.cell(row=i, column=mcol+1))
            Hyper_list=['Inbound',i,mcol+1]
#    print("edit inbound finished")

def edit_Outbound(j):
    mcol=Outbound.max_column
    for i in range(1,Outbound.max_row+1):
        Outbound.cell(row=i,column=mcol+1).value=Outbound.cell(row=i,column=mcol).value
        Outbound.cell(row=i,column=mcol+1)._style=copy(Outbound.cell(row=i,column=mcol)._style)
        if i==2:
            Outbound.cell(row=i,column=mcol+1).value=rep2ndrow(j.rstrip())
        elif i==4:
            Outbound.cell(row=i,column=mcol+1).value=j
            global Hyper_list
            Hyper_list=['Outbound',i,mcol+1]

def Hyperlink(r):
    global Hyper_list
    Index.cell(row=r, column=1).hyperlink= ("Quantum-II-MIM-IA.xlsx#{}!{}".format(Hyper_list[0],str(Index.cell(row=Hyper_list[1], column=Hyper_list[2])).replace('>','').split('.')[1]))
    Hyper_list=[] #Resetting the hyperlist

def edit_Index(j):
    global Vendor_sub
    Vendor_sub=[] #Resetting the vendor list
    lst=[]

    for r in range(1,Index.max_row+1):
        if re.search(j,'{}'.format(Index.cell(row=r,column=1).value)):
            lst.append(r)#this will give row names for multiple entries
    if lst:
        Index.cell(row=lst[0],column=3).value=PC_FILEID(j.rstrip())
        Index.cell(row=lst[0],column=4).value=Vendor_sub[0]
        Index.cell(row=lst[0],column=2).value=Vendor_sub[1]+"BOUND"
        Hyperlink(lst[0])
        lst.pop(0)#eliminate 1st occurance from consideration
        for d in lst[::-1]:  #delete all other rows
            Index.delete_rows(d,1)

    else:print("No entry found for: "+j)

with open(r"C:\Users\sahil.rana\Desktop\task\TSYS archiveR1.txt","r") as obj:

    print("1:Inbound")
    print("2:Outbound")
    User_Input=int(input("Make selection:"))
    for j in obj:
        if User_Input==1:
            edit_Inbound(j.rstrip())
        elif User_Input==2:
            edit_Outbound(j.rstrip())
        edit_Index(j.rstrip())

wb.save(r'C:\Users\sahil.rana\Desktop\task\Quantum-II-MIM-IA.xlsx')
