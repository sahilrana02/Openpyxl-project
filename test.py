import openpyxl as op
import re
from copy import copy
wb= op.load_workbook(r'C:\Users\RANA\Desktop\Book1.xlsx')
sheet=wb.get_sheet_by_name('Index')

def PC_FILEID(job):
    with open(r'C:\Users\RANA\Desktop\PFA\3 New\3 New\{}\MIM\SQL\insert-JOB-DEV-SYS-INT-PRD.sql'.format(job),'r') as obj:
        for i in obj:
            if re.search(",\d\d\d\d\);",i):
                break
    id=i.strip()
    id=id.replace(',','F'+id[4])
    id=id.replace('0);','')
    return id

with open(r"C:/Users\RANA\Desktop\PFA\R1 Jobs1.txt","r") as obj:
    for j in obj:
        lst=[]
        for r in range(1,sheet.max_row+1):
            if sheet.cell(row=r,column=1).value == j:
                lst.append(r)#this will give row names for multiple entries
                print(lst)
        if len(lst)>1:#incase there are more than one rows considering atleast one row fulfills the criteria
            for x in lst:
               if sheet.cell(row=x,column=3).value==PC_FILEID(j.rstrip()):
                   lst.remove(x)#eliminate correct row from consideration
            for w in lst[::-1]:
                sheet.delete_rows(w,1) #delete all other rows from last
                print(("lst>1"))
        elif len(lst)==1:           #incase only one entry found
            print("elif")
            sheet.cell(row=lst[0],column=3).value=PC_FILEID(j.rstrip())

wb.save(r'C:\Users\RANA\Desktop\Book1.xlsx')





