import openpyxl
import re
wb=openpyxl.load_workbook(r'C:\Users\RANA\Desktop\Book1.xlsx')
index=wb.get_sheet_by_name('Index')

index.cell(row=3, column=1).hyperlink = ("Book1.xlsx#sheet1!{}".format(str(index.cell(row=4, column=3)).replace('>','').split('.')[1]))

wb.save(r'C:\Users\RANA\Desktop\Book1.xlsx')
